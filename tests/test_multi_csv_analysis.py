from pathlib import Path
import sys

import pandas as pd
import pytest


PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from analytics import (
    build_analysis_tables,
    calculate_category_totals,
    calculate_month_totals,
    filter_expenses,
    prepare_expenses,
)
import analyze_expenses
from app_config import CLOUD_MODE, LOCAL_MODE, get_app_mode
import logging_config
from importer import combine_valid_expense_files, validate_expense_files
from importer import get_demo_file
from report import build_filtered_excel_report
from sections.compare_periods_section import (  # noqa: E402
    build_category_comparison,
    calculate_period_summary,
    get_comparison_values,
)
from validate_data import validate_csv


def write_csv(path, rows):
    path.write_text(
        "\n".join(["date,category,description,amount", *rows]),
        encoding="utf-8",
    )


def test_one_good_csv_passes_validation(tmp_path):
    csv_file = tmp_path / "good.csv"
    write_csv(
        csv_file,
        [
            "2026-01-01,Food,Breakfast,10.00",
            "2026-01-02,Transport,Bus,2.50",
        ],
    )

    assert validate_csv(csv_file) == []


def test_multiple_good_csv_files_are_combined_correctly(tmp_path):
    january = tmp_path / "january.csv"
    february = tmp_path / "february.csv"
    write_csv(january, ["2026-01-01,Food,Breakfast,10.00"])
    write_csv(february, ["2026-02-01,Health,Medicine,20.00"])

    expenses = combine_valid_expense_files([january, february])

    assert len(expenses) == 2
    assert expenses["amount"].sum() == 30.00


def test_month_column_is_added_after_combining_files(tmp_path):
    january = tmp_path / "january.csv"
    write_csv(january, ["2026-01-15,Food,Lunch,12.00"])

    expenses = combine_valid_expense_files([january])

    assert "month" in expenses.columns
    assert expenses.loc[0, "month"] == "2026-01"


def test_expenses_by_month_are_calculated_correctly(tmp_path):
    january = tmp_path / "january.csv"
    february = tmp_path / "february.csv"
    write_csv(
        january,
        [
            "2026-01-01,Food,Breakfast,10.00",
            "2026-01-02,Transport,Bus,5.00",
        ],
    )
    write_csv(february, ["2026-02-01,Health,Medicine,20.00"])

    expenses = combine_valid_expense_files([january, february])
    month_totals = calculate_month_totals(expenses)
    totals = dict(zip(month_totals["month"], month_totals["amount"]))

    assert totals == {"2026-01": 15.00, "2026-02": 20.00}


def test_expenses_by_category_are_calculated_correctly(tmp_path):
    csv_file = tmp_path / "good.csv"
    write_csv(
        csv_file,
        [
            "2026-01-01,Food,Breakfast,10.00",
            "2026-01-02,Food,Lunch,15.00",
            "2026-01-03,Transport,Bus,5.00",
        ],
    )

    expenses = combine_valid_expense_files([csv_file])
    category_totals = calculate_category_totals(expenses)
    totals = dict(zip(category_totals["category"], category_totals["amount"]))

    assert totals == {"Food": 25.00, "Transport": 5.00}


def test_bad_file_among_multiple_files_blocks_report_creation(tmp_path):
    good_file = tmp_path / "good.csv"
    bad_file = tmp_path / "bad.csv"
    report_file = tmp_path / "report.xlsx"
    write_csv(good_file, ["2026-01-01,Food,Breakfast,10.00"])
    write_csv(bad_file, ["2026-01-02,Food,Bad amount,not-a-number"])

    errors = validate_expense_files([good_file, bad_file])

    assert errors
    assert errors[0][0] == "bad.csv"
    assert not report_file.exists()


def test_empty_file_list_has_clear_error():
    with pytest.raises(ValueError, match="Please upload at least one CSV file"):
        validate_expense_files([])


def test_demo_csv_is_valid_and_can_be_analyzed():
    demo_file = get_demo_file()

    assert validate_csv(demo_file) == []

    expenses = combine_valid_expense_files([demo_file])

    assert len(expenses) > 0
    assert "month" in expenses.columns
    assert expenses["amount"].sum() > 0


def test_app_mode_defaults_to_local(monkeypatch):
    monkeypatch.delenv("APP_MODE", raising=False)

    assert get_app_mode() == LOCAL_MODE


def test_app_mode_can_be_set_to_cloud(monkeypatch):
    monkeypatch.setenv("APP_MODE", "cloud")

    assert get_app_mode() == CLOUD_MODE


def test_analysis_filters_and_tables_use_saved_expenses():
    expenses = prepare_expenses(
        [
            {
                "import_id": 1,
                "date": "2026-01-05",
                "category": "Food",
                "description": "Groceries",
                "amount": 30.00,
            },
            {
                "import_id": 1,
                "date": "2026-02-05",
                "category": "Transport",
                "description": "Bus ticket",
                "amount": 5.00,
            },
            {
                "import_id": 2,
                "date": "2026-02-10",
                "category": "Food",
                "description": "Cafe",
                "amount": 15.00,
            },
        ]
    )

    filtered = filter_expenses(
        expenses,
        start_date="2026-02-01",
        end_date="2026-02-28",
        categories=["Food"],
        import_ids=[2],
        description_text="cafe",
    )
    category_totals, month_totals, month_category_totals = build_analysis_tables(
        filtered
    )

    assert len(filtered) == 1
    assert filtered.iloc[0]["month"] == "2026-02"
    assert category_totals.iloc[0]["amount"] == 15.00
    assert month_totals.iloc[0]["amount"] == 15.00
    assert month_category_totals.loc["2026-02", "Food"] == 15.00


def filter_test_expenses():
    return prepare_expenses(
        [
            {
                "import_id": 1,
                "date": "2026-01-05",
                "category": "Food",
                "description": "Groceries",
                "amount": 30.00,
            },
            {
                "import_id": 2,
                "date": "2026-02-05",
                "category": "Transport",
                "description": "Bus ticket",
                "amount": 5.00,
            },
            {
                "import_id": 2,
                "date": "2026-02-10",
                "category": "Food",
                "description": "Cafe lunch",
                "amount": 15.00,
            },
        ]
    )


def test_filter_expenses_by_date_range():
    filtered = filter_expenses(
        filter_test_expenses(),
        start_date="2026-02-01",
        end_date="2026-02-28",
    )

    assert len(filtered) == 2
    assert filtered["month"].unique().tolist() == ["2026-02"]


def test_filter_expenses_by_category():
    filtered = filter_expenses(filter_test_expenses(), categories=["Food"])

    assert len(filtered) == 2
    assert filtered["category"].unique().tolist() == ["Food"]


def test_filter_expenses_by_description_text():
    filtered = filter_expenses(filter_test_expenses(), description_text="LUNCH")

    assert len(filtered) == 1
    assert filtered.iloc[0]["description"] == "Cafe lunch"


def test_period_comparison_calculates_categories_and_percentage_change():
    period_a = prepare_expenses(
        [
            {
                "import_id": 1,
                "date": "2026-01-05",
                "category": "Food",
                "description": "Groceries",
                "amount": 10.00,
            }
        ]
    )
    period_b = prepare_expenses(
        [
            {
                "import_id": 2,
                "date": "2026-02-05",
                "category": "Food",
                "description": "Cafe",
                "amount": 20.00,
            },
            {
                "import_id": 2,
                "date": "2026-02-07",
                "category": "Transport",
                "description": "Bus",
                "amount": 5.00,
            },
        ]
    )

    values = get_comparison_values(
        calculate_period_summary(period_a),
        calculate_period_summary(period_b),
    )
    categories = build_category_comparison(period_a, period_b).set_index("category")

    assert values["total_difference"] == 15.00
    assert values["percentage_change"] == 150.00
    assert values["average_difference"] == 2.50
    assert values["expense_count_difference"] == 1
    assert categories.loc["Food", "Period A"] == 10.00
    assert categories.loc["Transport", "Period B"] == 5.00


def test_period_comparison_handles_an_empty_period_without_division_by_zero():
    empty_period = prepare_expenses(
        pd.DataFrame(
            columns=["import_id", "date", "category", "description", "amount"]
        )
    )
    period_with_data = prepare_expenses(
        [
            {
                "import_id": 1,
                "date": "2026-01-05",
                "category": "Food",
                "description": "Groceries",
                "amount": 10.00,
            }
        ]
    )

    values = get_comparison_values(
        calculate_period_summary(empty_period),
        calculate_period_summary(period_with_data),
    )

    assert calculate_period_summary(empty_period) is None
    assert values["percentage_change"] is None


def test_filtered_excel_report_has_required_sheets_and_valid_checks():
    expenses = prepare_expenses(
        [
            {
                "import_id": 1,
                "date": "2026-01-05",
                "category": "Food",
                "description": "Groceries",
                "amount": 10.00,
            },
            {
                "import_id": 2,
                "date": "2026-02-05",
                "category": "Transport",
                "description": "Bus",
                "amount": 5.00,
            },
        ]
    )
    report_bytes = build_filtered_excel_report(
        expenses,
        {
            "date_range": "2026-01-01 to 2026-02-28",
            "categories": "All categories",
            "imports": "All imports",
            "description_text": "No text filter",
        },
    )

    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(report_bytes), data_only=True)
    summary = workbook["Summary"]
    validation = workbook["Validation"]

    assert workbook.sheetnames == [
        "Filtered Data",
        "Summary",
        "By Category",
        "By Month",
        "Month x Category",
        "Validation",
    ]
    assert summary["B7"].value == 2
    assert summary["B8"].value == 15.00
    assert [validation.cell(row, 3).value for row in range(2, 5)] == [
        "OK",
        "OK",
        "OK",
    ]


def test_command_line_mode_still_creates_report_for_one_csv(tmp_path, monkeypatch):
    output_file = tmp_path / "single_csv_report.xlsx"
    log_file = tmp_path / "run_log.txt"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "analyze_expenses.py",
            "--input",
            str(PROJECT_ROOT / "data" / "expenses.csv"),
            "--output",
            str(output_file),
        ],
    )
    monkeypatch.setattr(logging_config, "LOG_FILE", log_file)
    monkeypatch.setattr(analyze_expenses, "wait_before_close", lambda: None)

    analyze_expenses.main()

    assert output_file.exists()
    assert log_file.exists()
